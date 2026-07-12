import os

from openpyxl import load_workbook

from src.analytics.peer import PEER_OUTPUT


def test_peer_report_exists():

    assert os.path.exists(
        PEER_OUTPUT
    )


def test_workbook_has_sheets():

    wb = load_workbook(
        PEER_OUTPUT
    )

    assert len(
        wb.sheetnames
    ) == 11


def test_private_banks_sheet_exists():

    wb = load_workbook(
        PEER_OUTPUT
    )

    assert (
        "Private Banks"
        in wb.sheetnames
    )


def test_benchmark_row_exists():

    wb = load_workbook(
        PEER_OUTPUT
    )

    ws = wb["Private Banks"]

    found = False

    for row in ws.iter_rows():

        for cell in row:

            if cell.value == 1:

                found = True

                break

    assert found


def test_median_row_exists():

    wb = load_workbook(
        PEER_OUTPUT
    )

    ws = wb["Private Banks"]

    last_row = ws.max_row

    assert (
        ws.cell(
            row=last_row,
            column=1,
        ).value
        == "Median"
    )