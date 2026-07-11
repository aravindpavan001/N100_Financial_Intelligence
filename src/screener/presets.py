import yaml
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.screener.engine import (
    apply_filters,
    load_financial_ratios,
    DB_PATH,
)

from src.analytics.score import (
    calculate_profitability_score,
    calculate_cash_quality_score,
    calculate_growth_score,
    calculate_leverage_score,
    calculate_composite_score,
)


PRESET_PATH = "config/presets.yaml"

OUTPUT_FILE = os.path.join(
    "output",
    "screener_output.xlsx",
)


def load_presets(path):

    with open(
        path,
        "r",
        encoding="utf-8",
    ) as file:

        presets = yaml.safe_load(file)

    return presets


def run_preset(
    df,
    preset_name,
    presets,
):

    if preset_name not in presets:
        raise KeyError(
            f"Preset '{preset_name}' not found."
        )

    filters = {
        "filters": presets[preset_name]
    }

    result = apply_filters(
        df,
        filters,
    )

    if "composite_quality_score" in result.columns:

     result = result.sort_values(
        "composite_quality_score",
        ascending=False,
    )

    return result


if __name__ == "__main__":

    presets = load_presets(
        PRESET_PATH,
    )

    df = load_financial_ratios(
        DB_PATH,
    )
    df = calculate_profitability_score(df)

    df = calculate_cash_quality_score(df)

    df = calculate_growth_score(df)

    df = calculate_leverage_score(df)

    df = calculate_composite_score(df)
     
     
    writer = pd.ExcelWriter(
      OUTPUT_FILE,
      engine="openpyxl",
     )
    print("=" * 70)
    print("PRESET RESULTS")
    print("=" * 70)

    for preset_name in presets:

        result = run_preset(
            df,
            preset_name,
            presets,
        )

        result.to_excel(
            writer,
            sheet_name=preset_name[:31],
            index=False,
        )

        print()

        print(
            result[
                [
                    "company_id",
                    "composite_quality_score",
                    "return_on_equity_pct",
                    "revenue_cagr_5yr",
                ]
            ].head()
        )

        print("-" * 70)

        print(
            f"{preset_name:<25} : {len(result)} companies"
        )

        
    writer.close()

    wb = load_workbook(
        OUTPUT_FILE,
    )

    green = PatternFill(
        fill_type="solid",
        fgColor="90EE90",
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for sheet in wb.worksheets:

        headers = {}

        for cell in sheet[1]:
            headers[cell.value] = cell.column

        if "return_on_equity_pct" in headers:

            roe_col = headers["return_on_equity_pct"]

            for row in range(
                2,
                sheet.max_row + 1,
            ):

                cell = sheet.cell(
                    row=row,
                    column=roe_col,
                )

                if (
                    cell.value is not None
                    and cell.value >= 15
                ):
                    cell.fill = green
                else:
                    cell.fill = red

        if "debt_to_equity" in headers:

            de_col = headers["debt_to_equity"]

            for row in range(
                2,
                sheet.max_row + 1,
            ):

                cell = sheet.cell(
                    row=row,
                    column=de_col,
                )

                if (
                    cell.value is not None
                    and cell.value <= 1
                ):
                    cell.fill = green
                else:
                    cell.fill = red           

    wb.save(
        OUTPUT_FILE,
    )

    print()
    print("Excel Generated:")
    print(OUTPUT_FILE)