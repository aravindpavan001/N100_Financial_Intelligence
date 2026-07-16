import os
import sqlite3

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl.styles import PatternFill
from openpyxl import load_workbook


# =====================================================
# PATHS
# =====================================================

DB_PATH = "nifty100.db"

RADAR_FOLDER = os.path.join(
    "reports",
    "radar_charts",
)

PEER_OUTPUT = os.path.join(
    "output",
    "peer_comparison.xlsx",
)

os.makedirs(
    RADAR_FOLDER,
    exist_ok=True,
)


# =====================================================
# RADAR METRICS
# =====================================================

RADAR_METRICS = [

    "return_on_equity_pct",

    "operating_profit_margin_pct",

    "net_profit_margin_pct",

    "debt_to_equity",

    "free_cash_flow_cr",

    "pat_cagr_5yr",

    "revenue_cagr_5yr",

    "composite_quality_score",

]


# =====================================================
# LOAD SQLITE TABLES
# =====================================================

def load_peer_groups(db_path):

    conn = sqlite3.connect(db_path)

    df = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn,
    )

    conn.close()

    return df


def load_financial_ratios(db_path):

    conn = sqlite3.connect(db_path)

    df = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn,
    )

    conn.close()

    return df


def load_peer_percentiles(db_path):

    conn = sqlite3.connect(db_path)

    df = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn,
    )

    conn.close()

    return df

def prepare_peer_report_data(
    db_path,
):

    peer_df = load_peer_groups(
        db_path,
    )

    ratio_df = load_financial_ratios(
        db_path,
    )

    percentile_df = load_peer_percentiles(
        db_path,
    )

    merged = peer_df.merge(
        ratio_df,
        on="company_id",
        how="left",
    )

    merged = merged.merge(
        percentile_df,
        on=[
            "company_id",
            "year",
        ],
        how="left",
    )

    return merged

def pivot_percentiles(
    percentile_df,
):

    pivot = percentile_df.pivot_table(
        index=[
            "company_id",
            "peer_group_name",
            "year",
        ],
        columns="metric",
        values="percentile_rank",
    )

    pivot = pivot.reset_index()

    return pivot

def create_final_peer_report(
    db_path,
):

    peer_df = load_peer_groups(
        db_path,
    )

    ratio_df = load_financial_ratios(
        db_path,
    )

    percentile_df = load_peer_percentiles(
        db_path,
    )

    pivot_df = pivot_percentiles(
        percentile_df,
    )

    report_df = peer_df.merge(
        ratio_df,
        on="company_id",
        how="left",
    )

    report_df = report_df.merge(
        pivot_df,
        on=[
            "company_id",
            "peer_group_name",
            "year",
        ],
        how="left",
    )
    
    report_df = report_df.sort_values(
    "year"
)

    report_df = report_df.groupby(
    "company_id",
    as_index=False,
    ).tail(1)
    
    return report_df



def export_peer_report(
    final_df,
):

    writer = pd.ExcelWriter(
        PEER_OUTPUT,
        engine="openpyxl",
    )

    peer_groups = (
        final_df["peer_group_name"]
        .dropna()
        .unique()
    )

    for group in peer_groups:

        group_df = final_df[
            final_df["peer_group_name"] == group
        ]

        group_df.to_excel(
            writer,
            sheet_name=group[:31],
            index=False,
        )

    writer.close()

    wb = load_workbook(
        PEER_OUTPUT,
    )

    gold = PatternFill(
        fill_type="solid",
        fgColor="FFD966",
    )

    green = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        benchmark_col = headers.index(
            "is_benchmark"
        ) + 1

        percentile_columns = []

        for i, header in enumerate(
            headers,
            start=1,
        ):

            if (
                header is not None
                and header.endswith("_y")
            ):

                percentile_columns.append(i)

        # -------------------------
        # Gold Benchmark Row
        # -------------------------

        for row in range(
            2,
            ws.max_row + 1,
        ):

            if ws.cell(
                row=row,
                column=benchmark_col,
            ).value == 1:

                for col in range(
                    1,
                    ws.max_column + 1,
                ):

                    ws.cell(
                        row=row,
                        column=col,
                    ).fill = gold

        # -------------------------
        # Percentile Colours
        # -------------------------

        for col in percentile_columns:

            for row in range(
                2,
                ws.max_row + 1,
            ):

                cell = ws.cell(
                    row=row,
                    column=col,
                )

                value = cell.value

                if (
                    value is None
                    or not isinstance(
                        value,
                        (
                            int,
                            float,
                        ),
                    )
                ):
                    continue

                if value >= 75:

                    cell.fill = green

                elif value >= 25:

                    cell.fill = yellow

                else:

                    cell.fill = red

        # -------------------------
        # Median Row
        # -------------------------

        median_row = ws.max_row + 1

        ws.cell(
            row=median_row,
            column=1,
        ).value = "Median"

        for col in range(
            2,
            ws.max_column + 1,
        ):

            values = []

            for row in range(
                2,
                median_row,
            ):

                value = ws.cell(
                    row=row,
                    column=col,
                ).value

                if isinstance(
                    value,
                    (
                        int,
                        float,
                    ),
                ):

                    values.append(value)

            if values:

                values.sort()

                n = len(values)

                if n % 2 == 0:

                    median = (
                        values[n // 2 - 1]
                        + values[n // 2]
                    ) / 2

                else:

                    median = values[n // 2]

                ws.cell(
                    row=median_row,
                    column=col,
                ).value = round(
                    median,
                    2,
                )

    wb.save(
        PEER_OUTPUT,
    )

    print()

    print("Peer comparison workbook created.")
# =====================================================
# MERGE
# =====================================================

def merge_peer_data(
    peer_df,
    ratio_df,
):

    merged = peer_df.merge(

        ratio_df,

        on="company_id",

        how="left",

    )

    return merged


# =====================================================
# PERCENTILE RANK
# =====================================================

def calculate_percentile_rank(

    df,

    metric,

    higher_is_better=True,

):

    if higher_is_better:

        rank = df[metric].rank(

            method="average",

            pct=True,

            ascending=True,

        )

    else:

        rank = df[metric].rank(

            method="average",

            pct=True,

            ascending=False,

        )

    return rank * 100


METRICS = {

    "return_on_equity_pct": True,

    "operating_profit_margin_pct": True,

    "net_profit_margin_pct": True,

    "debt_to_equity": False,

    "free_cash_flow_cr": True,

    "revenue_cagr_5yr": True,

    "pat_cagr_5yr": True,

    "eps_cagr_5yr": True,

    "interest_coverage": True,

    "asset_turnover": True,

}
# =====================================================
# CALCULATE ALL PEER PERCENTILES
# =====================================================

def calculate_all_percentiles(
    merged_df,
):

    results = []

    peer_groups = (
        merged_df["peer_group_name"]
        .dropna()
        .unique()
    )

    for peer_group in peer_groups:

        peer_data = merged_df[
            merged_df["peer_group_name"]
            == peer_group
        ]

        years = (
            peer_data["year"]
            .dropna()
            .unique()
        )

        for year in years:

            year_data = peer_data[
                peer_data["year"] == year
            ]

            for metric, higher_is_better in METRICS.items():

                if metric not in year_data.columns:
                    continue

                temp = year_data.copy()

                temp["percentile_rank"] = (
                    calculate_percentile_rank(
                        temp,
                        metric,
                        higher_is_better,
                    )
                )

                temp["metric"] = metric

                temp["metric_value"] = temp[metric]

                results.append(

                    temp[
                        [
                            "company_id",
                            "peer_group_name",
                            "year",
                            "metric",
                            "metric_value",
                            "percentile_rank",
                        ]
                    ]

                )

    final_df = pd.concat(
        results,
        ignore_index=True,
    )

    return final_df


# =====================================================
# SAVE TO SQLITE
# =====================================================

def save_peer_percentiles(
    df,
    db_path,
):

    conn = sqlite3.connect(
        db_path,
    )

    cursor = conn.cursor()

    cursor.execute(
        """
        DROP TABLE IF EXISTS peer_percentiles
        """
    )

    cursor.execute(
        """
        CREATE TABLE peer_percentiles (

            company_id TEXT,

            peer_group_name TEXT,

            year TEXT,

            metric TEXT,

            metric_value REAL,

            percentile_rank REAL

        )
        """
    )

    df.to_sql(
        "peer_percentiles",
        conn,
        if_exists="append",
        index=False,
    )

    conn.commit()

    conn.close()


# =====================================================
# COMPANY METRICS
# =====================================================

def get_company_metrics(
    merged_df,
    company_id,
):

    company = merged_df[
        merged_df["company_id"] == company_id
    ]

    if company.empty:
        return None

    company = company.sort_values(
        "year"
    )

    latest = company.iloc[-1]

    values = []

    for metric in RADAR_METRICS:

        if metric in latest.index:
            values.append(
                latest[metric]
            )
        else:
            values.append(
                np.nan
            )

    return values


# =====================================================
# PEER AVERAGE
# =====================================================

def get_peer_average(
    merged_df,
    peer_group,
):

    peer = merged_df[
        merged_df["peer_group_name"]
        == peer_group
    ]

    averages = []

    for metric in RADAR_METRICS:

        if metric in peer.columns:

            averages.append(
                peer[metric].mean()
            )

        else:

            averages.append(
                np.nan
            )

    return averages

def get_nifty_average(
    merged_df,
):

    averages = []

    for metric in RADAR_METRICS:

        if metric in merged_df.columns:

            averages.append(
                merged_df[metric].mean()
            )

        else:

            averages.append(np.nan)

    return averages    

# =====================================================
# RADAR CHART
# =====================================================

def generate_radar_chart(
    company_values,
    peer_values,
    labels,
    company_name,
    output_path,
):

    company_values = [
        0 if pd.isna(x) else x
        for x in company_values
    ]

    peer_values = [
        0 if pd.isna(x) else x
        for x in peer_values
    ]

    num_vars = len(labels)

    angles = np.linspace(
        0,
        2 * np.pi,
        num_vars,
        endpoint=False,
    ).tolist()

    company_values += company_values[:1]
    peer_values += peer_values[:1]
    angles += angles[:1]

    fig, ax = plt.subplots(
        figsize=(8, 8),
        subplot_kw=dict(
            polar=True,
        ),
    )

    # Company

    ax.plot(
        angles,
        company_values,
        linewidth=2,
        color="blue",
        label=company_name,
    )

    ax.fill(
        angles,
        company_values,
        alpha=0.25,
        color="blue",
    )

    # Peer Average

    ax.plot(
        angles,
        peer_values,
        linewidth=2,
        linestyle="--",
        color="red",
        label="Reference Average",
    )

    ax.set_xticks(
        angles[:-1]
    )

    ax.set_xticklabels(
        labels,
        fontsize=10,
    )

    ax.set_title(
        company_name,
        fontsize=14,
        pad=20,
    )

    ax.legend(
        loc="upper right",
    )

    plt.tight_layout()

    plt.savefig(
        output_path,
        dpi=300,
    )

    plt.close()


# =====================================================
# GENERATE ALL RADAR CHARTS
# =====================================================

def generate_all_radar_charts(
    merged_df,
):

    labels = [
        "ROE",
        "OPM",
        "NPM",
        "D/E",
        "FCF",
        "PAT CAGR",
        "Revenue CAGR",
        "Composite",
    ]

    companies = (
        merged_df["company_id"]
        .dropna()
        .unique()
    )

    created = 0
    skipped = 0

    for company_id in companies:

        company_rows = merged_df[
            merged_df["company_id"] == company_id
        ]

        if company_rows.empty:
            skipped += 1
            continue

        peer_group = company_rows.iloc[-1]["peer_group_name"]

        company_values = get_company_metrics(
            merged_df,
            company_id,
        )

        if pd.isna(peer_group):

            peer_values = get_nifty_average(
                merged_df,
            )

        else:

            peer_values = get_peer_average(
                merged_df,
                peer_group,
            )

        generate_radar_chart(
            company_values,
            peer_values,
            labels,
            company_id,
            os.path.join(
                RADAR_FOLDER,
                f"{company_id}_radar.png",
            ),
        )

        created += 1

    print()
    print("=" * 50)
    print("RADAR CHART SUMMARY")
    print("=" * 50)
    print(f"Charts Created : {created}")
    print(f"Charts Skipped : {skipped}")
    print(f"Folder         : {RADAR_FOLDER}")
    print("=" * 50)
    print()

    # =====================================================
# MAIN
# =====================================================

if __name__ == "__main__":

    print("=" * 60)
    print("LOADING TABLES")
    print("=" * 60)

    peer_df = load_peer_groups(
        DB_PATH,
    )

    ratio_df = load_financial_ratios(
        DB_PATH,
    )

    print()

    print("Peer Groups Shape")
    print(peer_df.shape)

    print()

    print("Financial Ratios Shape")
    print(ratio_df.shape)

    print()

    print("=" * 60)
    print("MERGING DATA")
    print("=" * 60)

    merged_df = merge_peer_data(
        peer_df,
        ratio_df,
    )

    print()

    print("Merged Shape")
    print(merged_df.shape)

    print()

    print("=" * 60)
    print("CALCULATING PEER PERCENTILES")
    print("=" * 60)

    percentile_df = calculate_all_percentiles(
        merged_df,
    )

    print()

    print("Percentile Rows")
    print(len(percentile_df))

    print()

    print(percentile_df.head())

    print()

    print("=" * 60)
    print("SAVING SQLITE TABLE")
    print("=" * 60)

    save_peer_percentiles(
        percentile_df,
        DB_PATH,
    )

    print()

    print("peer_percentiles table saved.")

    print()

    print("=" * 60)
    print("GENERATING RADAR CHARTS")
    print("=" * 60)

    generate_all_radar_charts(
        merged_df,
    )

    print()

    print("=" * 60)
    print("DAY 19 COMPLETED")
    print("=" * 60)

    print("✔ Peer Percentiles Generated")

    print("✔ SQLite Updated")

    print("✔ Radar Charts Generated")

    print()

    print("Charts Folder")

    print(RADAR_FOLDER)

    print()

    print("=" * 60)
    print("DAY 20")
    print("=" * 60)

    final_df = create_final_peer_report(
    DB_PATH,
    )

    export_peer_report(
    final_df,
    )

    print()

    print("Peer comparison workbook created.")

    print()

    print(final_df.head())

    print()

    print(final_df.shape)

    print()

    print("=" * 60)

    print()



